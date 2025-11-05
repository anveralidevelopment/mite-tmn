import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import requests
from bs4 import BeautifulSoup
from datetime import datetime, date, timedelta
import re
import json
import os
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from fake_useragent import UserAgent
import threading
import feedparser
from tkcalendar import DateEntry
import pygame
import sys
from dateutil import parser as date_parser
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import time
from logger_config import setup_logger

class TickActivityMonitor:
    def __init__(self, root):
        # Инициализация логирования (до загрузки конфига)
        self.logger = setup_logger()
        self.logger.info("Инициализация приложения")
        
        # Загрузка конфигурации
        self.config = self._load_config()
        
        self.root = root
        self.root.title(self.config.get('app', {}).get('title', 'Монитор активности клещей - Тюмень'))
        self.root.geometry(self.config.get('app', {}).get('geometry', '1200x800'))
        
        # Инициализация звука
        pygame.mixer.init()
        sound_config = self.config.get('ui', {}).get('sounds', {})
        self.sound_enabled = sound_config.get('enabled', True)
        self.load_sounds()
        
        # Цветовая схема из конфига
        colors = self.config.get('ui', {}).get('colors', {})
        self.bg_color = colors.get('bg', '#0a0a1a')
        self.text_color = colors.get('text', '#00ff99')
        self.highlight_color = colors.get('highlight', '#ff66ff')
        self.button_color = colors.get('button', '#1a3a3a')
        self.button_active_color = colors.get('button_active', '#2a5a5a')
        
        # Настройка стилей
        self.setup_styles()
        
        # Создание главного холста
        self.canvas = tk.Canvas(root, highlightthickness=0, bg=self.bg_color)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        
        # Неоновая обводка (толще)
        self.neon_border = None
        self.create_neon_border()
        self.animate_neon_border()
        
        # Данные из конфига
        app_config = self.config.get('app', {})
        self.data_file = app_config.get('data_file', 'tick_data.json')
        self.excel_file = app_config.get('excel_file', 'tick_stats.xlsx')
        self.default_data = {
            "current_week": {"cases": 0, "risk_level": "Нет данных", "date": ""},
            "previous_week": {"cases": 0, "risk_level": "Нет данных", "date": ""},
            "sources": [],
            "last_update": "никогда"
        }
        self.data = self.default_data.copy()
        
        # Инициализация Excel файла
        self.init_excel()
        
        # Создание интерфейса
        self.create_widgets()
        self.load_data()
        
        # Обработка изменения размера окна
        self.root.bind("<Configure>", self.on_window_resize)
        
        # Воспроизведение фоновой музыки
        self.play_background_sound()

    def init_excel(self):
        """Инициализация Excel файла с нужными колонками"""
        if not os.path.exists(self.excel_file):
            df = pd.DataFrame(columns=[
                "Дата", "Количество случаев", "Уровень риска", 
                "Источник", "Заголовок", "Ссылка"
            ])
            df.to_excel(self.excel_file, index=False, engine='openpyxl')
            
            # Добавляем форматирование
            self.format_excel_file()

    def format_excel_file(self):
        """Добавляем форматирование в Excel файл"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            # Форматирование заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                ws.column_dimensions[get_column_letter(col)].width = 20
                
            # Автофильтр
            ws.auto_filter.ref = ws.dimensions
            
            wb.save(self.excel_file)
            wb.close()
        except Exception as e:
            self.logger.error(f"Ошибка при форматировании Excel: {str(e)}")

    def save_to_excel(self, data):
        """Сохранение данных в Excel файл"""
        try:
            # Читаем существующие данные
            df_existing = pd.read_excel(self.excel_file, engine='openpyxl')
            
            # Создаем DataFrame из новых данных
            new_rows = []
            for item in data:
                new_rows.append({
                    "Дата": item['date'].strftime('%d.%m.%Y') if isinstance(item['date'], date) else item['date'],
                    "Количество случаев": item['cases'],
                    "Уровень риска": self.calculate_risk_level(item['cases']),
                    "Источник": item.get('source', 'Неизвестно'),
                    "Заголовок": item.get('title', ''),
                    "Ссылка": item.get('url', '')
                })
            
            df_new = pd.DataFrame(new_rows)
            
            # Объединяем и удаляем дубликаты
            df_combined = pd.concat([df_existing, df_new])
            df_combined.drop_duplicates(
                subset=["Дата", "Источник", "Заголовок"], 
                keep='last', 
                inplace=True
            )
            
            # Сохраняем обратно в Excel
            df_combined.to_excel(self.excel_file, index=False, engine='openpyxl')
            
            # Применяем форматирование
            self.format_excel_file()
            
            self.logger.info(f"Данные сохранены в Excel: {len(new_rows)} новых записей")
            return True
        except Exception as e:
            self.logger.error(f"Ошибка при сохранении в Excel: {str(e)}")
            return False

    def load_from_excel(self):
        """Загрузка данных из Excel файла"""
        try:
            df = pd.read_excel(self.excel_file, engine='openpyxl')
            
            # Преобразуем даты из строк в объекты date
            df['Дата'] = pd.to_datetime(df['Дата'], format='%d.%m.%Y').dt.date
            
            # Сортируем по дате
            df = df.sort_values(by='Дата', ascending=False)
            
            # Конвертируем в список словарей
            sources = []
            for _, row in df.iterrows():
                sources.append({
                    'date': row['Дата'],
                    'cases': row['Количество случаев'],
                    'risk_level': row['Уровень риска'],
                    'source': row['Источник'],
                    'title': row['Заголовок'],
                    'url': row['Ссылка']
                })
            
            self.logger.info(f"Загружено {len(sources)} записей из Excel")
            return sources
        except Exception as e:
            self.logger.error(f"Ошибка при загрузке из Excel: {str(e)}")
            return []

    def _load_config(self):
        """Загрузка конфигурации из файла"""
        try:
            if os.path.exists('config.json'):
                with open('config.json', 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.logger.info("Конфигурация загружена из config.json")
                    return config
            else:
                self.logger.warning("Файл config.json не найден, используются значения по умолчанию")
                return {}
        except Exception as e:
            self.logger.error(f"Ошибка загрузки конфигурации: {str(e)}")
            return {}
    
    def load_sounds(self):
        """Загрузка звуковых эффектов"""
        try:
            self.background_sound = pygame.mixer.Sound("forest.wav") if os.path.exists("forest.wav") else None
            self.button_sound = pygame.mixer.Sound("button.wav") if os.path.exists("button.wav") else None
            self.update_sound = pygame.mixer.Sound("update.wav") if os.path.exists("update.wav") else None
            self.export_sound = pygame.mixer.Sound("export.wav") if os.path.exists("export.wav") else None
            self.logger.info("Звуковые эффекты загружены")
        except Exception as e:
            self.logger.error(f"Ошибка загрузки звуков: {str(e)}")
            self.sound_enabled = False

    def play_background_sound(self):
        """Воспроизведение фонового звука"""
        if self.sound_enabled and self.background_sound:
            volume = self.config.get('ui', {}).get('sounds', {}).get('background_volume', 0.3)
            self.background_sound.set_volume(volume)
            self.background_sound.play(loops=-1)

    def play_sound(self, sound):
        """Воспроизведение звукового эффекта"""
        if self.sound_enabled and sound:
            volume = self.config.get('ui', {}).get('sounds', {}).get('effect_volume', 0.5)
            sound.set_volume(volume)
            sound.play()

    def create_neon_border(self):
        """Создание неоновой обводки (толще)"""
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        
        if width > 10 and height > 10:
            if hasattr(self, 'neon_border'):
                self.canvas.delete(self.neon_border)
            
            self.neon_border = self.canvas.create_rectangle(
                5, 5, width-5, height-5, 
                outline="#00ffff", width=5,
                tags=("neon_border",)
            )

    def animate_neon_border(self):
        """Анимация неоновой обводки"""
        colors = ["#00ffff", "#ff00ff", "#00ff99", "#ff6600", "#ffffff"]
        current_color = self.canvas.itemcget(self.neon_border, "outline") if self.neon_border else colors[0]
        next_color = colors[(colors.index(current_color) + 1) % len(colors)] if current_color in colors else colors[0]
        
        if self.neon_border:
            self.canvas.itemconfig(self.neon_border, outline=next_color)
        self.root.after(300, self.animate_neon_border)

    def on_window_resize(self, event):
        """Обработка изменения размера окна"""
        self.create_neon_border()

    def setup_styles(self):
        """Настройка стилей интерфейса"""
        self.style = ttk.Style()
        self.style.theme_use('alt')
        self.root.configure(bg=self.bg_color)
        
        self.style.configure('.', 
                           background=self.bg_color, 
                           foreground=self.text_color,
                           font=('Arial', 10))
        
        self.style.configure('TFrame', background=self.bg_color)
        
        self.style.configure('TButton', 
                           background=self.button_color, 
                           foreground=self.text_color,
                           font=('Arial', 10, 'bold'),
                           borderwidth=1)
        self.style.map('TButton', 
                      background=[('active', self.button_active_color)],
                      foreground=[('active', '#ffffff')])
        
        self.style.configure('TNotebook', background=self.bg_color)
        self.style.configure('TNotebook.Tab', 
                           background="#1a1a3a",
                           foreground=self.text_color,
                           padding=[10, 5],
                           font=('Arial', 10, 'bold'))
        self.style.map('TNotebook.Tab', 
                      background=[('selected', self.button_color)],
                      foreground=[('selected', self.highlight_color)])
        
        self.style.configure('Header.TLabel', 
                           font=('Arial', 16, 'bold'), 
                           foreground=self.highlight_color)
        
        self.style.configure('Data.TLabel', 
                           font=('Arial', 12, 'bold'), 
                           foreground='#00ff99')
        
        self.style.configure('Risk.TLabel', 
                           font=('Arial', 14, 'bold'))
        
        self.style.configure('DateEntry', 
                           fieldbackground=self.bg_color,
                           foreground=self.text_color,
                           background=self.button_color)

    def create_widgets(self):
        """Создание элементов интерфейса"""
        main_frame = ttk.Frame(self.canvas)
        self.canvas.create_window((10, 10), window=main_frame, anchor="nw", tags=("main_frame",))
        
        self.create_filter_panel(main_frame)
        
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        self.create_stats_tab()
        self.create_sources_tab()
        self.create_control_panel()

    def create_filter_panel(self, parent):
        """Создание панели фильтров"""
        filter_frame = ttk.LabelFrame(parent, text="Фильтр данных", style='TFrame')
        filter_frame.pack(fill=tk.X, padx=10, pady=10)

        ttk.Label(filter_frame, text="Начальная дата:").grid(row=0, column=0, padx=5, pady=5)
        self.start_date = DateEntry(filter_frame, 
                                  date_pattern='dd.mm.yyyy',
                                  background='darkblue', 
                                  foreground='white', 
                                  borderwidth=2)
        self.start_date.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(filter_frame, text="Конечная дата:").grid(row=0, column=2, padx=5, pady=5)
        self.end_date = DateEntry(filter_frame, 
                                date_pattern='dd.mm.yyyy',
                                background='darkblue', 
                                foreground='white', 
                                borderwidth=2)
        self.end_date.grid(row=0, column=3, padx=5, pady=5)

        apply_btn = ttk.Button(filter_frame, 
                             text="Применить фильтр", 
                             command=self.apply_date_filter)
        apply_btn.grid(row=0, column=4, padx=10, pady=5)
        apply_btn.bind("<Button-1>", lambda e: self.play_sound(self.button_sound))

        export_btn = ttk.Button(filter_frame, 
                              text="📁 Экспорт в Excel", 
                              command=self.export_to_excel)
        export_btn.grid(row=0, column=5, padx=10, pady=5)
        export_btn.bind("<Button-1>", lambda e: self.play_sound(self.button_sound))

    def export_to_excel(self):
        """Экспорт данных в новый Excel файл"""
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile="tick_stats_export.xlsx"
            )
            
            if file_path:
                # Копируем наш основной файл
                import shutil
                shutil.copyfile(self.excel_file, file_path)
                
                # Форматируем экспортированный файл
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                
                # Добавляем заголовок
                ws.insert_rows(1)
                ws.merge_cells('A1:F1')
                title_cell = ws['A1']
                title_cell.value = "Данные по активности клещей в Тюменской области"
                title_cell.font = Font(bold=True, size=14)
                title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                
                # Добавляем дату экспорта
                ws.insert_rows(2)
                ws.merge_cells('A2:F2')
                date_cell = ws['A2']
                date_cell.value = f"Дата экспорта: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
                date_cell.font = Font(italic=True)
                date_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                
                wb.save(file_path)
                wb.close()
                
                messagebox.showinfo("Успех", f"Данные успешно экспортированы в файл:\n{file_path}")
                self.play_sound(self.export_sound)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось экспортировать данные: {str(e)}")

    def create_stats_tab(self):
        """Создание вкладки со статистикой"""
        stats_tab = ttk.Frame(self.notebook)
        self.notebook.add(stats_tab, text="📊 Статистика активности")

        header = ttk.Label(stats_tab, 
                         text="АКТИВНОСТЬ КЛЕЩЕЙ В ТЮМЕНСКОЙ ОБЛАСТИ", 
                         style='Header.TLabel')
        header.pack(pady=10)

        data_frame = ttk.Frame(stats_tab)
        data_frame.pack(fill=tk.X, pady=10)

        ttk.Label(data_frame, text="Текущая неделя:").grid(row=0, column=0, padx=20, sticky=tk.W)
        self.current_data = ttk.Label(data_frame, text="0 случаев\n(дата неизвестна)", style='Data.TLabel')
        self.current_data.grid(row=1, column=0, padx=20)
        self.current_risk = ttk.Label(data_frame, text="Риск: Нет данных", style='Risk.TLabel')
        self.current_risk.grid(row=2, column=0, padx=20)

        ttk.Label(data_frame, text="Прошлая неделя:").grid(row=0, column=1, padx=20, sticky=tk.W)
        self.prev_data = ttk.Label(data_frame, text="0 случаев\n(дата неизвестна)", style='Data.TLabel')
        self.prev_data.grid(row=1, column=1, padx=20)
        self.prev_risk = ttk.Label(data_frame, text="Риск: Нет данных", style='Risk.TLabel')
        self.prev_risk.grid(row=2, column=1, padx=20)

        self.setup_graph(stats_tab)

    def setup_graph(self, parent):
        """Настройка графика"""
        graph_frame = ttk.Frame(parent)
        graph_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self.figure = plt.Figure(figsize=(10, 5), dpi=100, facecolor=self.bg_color)
        self.ax = self.figure.add_subplot(111)
        self.ax.set_facecolor(self.bg_color)
        self.ax.tick_params(colors=self.text_color, labelsize=8)
        
        # Настройка осей
        for spine in self.ax.spines.values():
            spine.set_color(self.text_color)
            spine.set_linewidth(1.5)
        
        # Настройка сетки
        self.ax.grid(True, linestyle='--', alpha=0.7, color='#444444')
        
        self.canvas_graph = FigureCanvasTkAgg(self.figure, graph_frame)
        self.canvas_graph.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def create_sources_tab(self):
        """Создание вкладки с источниками"""
        sources_tab = ttk.Frame(self.notebook)
        self.notebook.add(sources_tab, text="📰 Источники данных")
        
        self.sources_text = tk.Text(sources_tab, 
                                  wrap=tk.WORD, 
                                  bg="#1a3a3a",
                                  fg=self.text_color,
                                  font=('Arial', 10), 
                                  padx=10,
                                  pady=10,
                                  insertbackground=self.text_color,
                                  selectbackground=self.highlight_color)
        self.sources_text.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(self.sources_text)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.sources_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.sources_text.yview)

    def create_control_panel(self):
        """Создание панели управления"""
        control_frame = ttk.Frame(self.canvas)
        self.canvas.create_window((10, 10), window=control_frame, anchor="sw", tags=("control_frame",))
        control_frame.pack(fill=tk.X, pady=5)
        
        self.update_btn = ttk.Button(control_frame, 
                                   text="🔄 Обновить данные", 
                                   command=self.update_data_threaded)
        self.update_btn.pack(side=tk.LEFT, padx=5)
        
        self.sound_btn = ttk.Button(control_frame, 
                                   text="🔊 Звук Вкл/Выкл", 
                                   command=self.toggle_sound)
        self.sound_btn.pack(side=tk.LEFT, padx=5)
        
        self.risk_legend_btn = ttk.Button(control_frame,
                                        text="📊 Легенда рисков",
                                        command=self.show_risk_legend)
        self.risk_legend_btn.pack(side=tk.LEFT, padx=5)
        
        self.update_label = ttk.Label(control_frame, style='Data.TLabel')
        self.update_label.pack(side=tk.RIGHT, padx=10)
        
        self.update_btn.bind("<Button-1>", lambda e: self.play_sound(self.button_sound))
        self.sound_btn.bind("<Button-1>", lambda e: self.play_sound(self.button_sound))
        self.risk_legend_btn.bind("<Button-1>", lambda e: self.play_sound(self.button_sound))

    def show_risk_legend(self):
        """Показывает легенду уровней риска"""
        legend_window = tk.Toplevel(self.root)
        legend_window.title("Легенда уровней риска")
        legend_window.geometry("300x200")
        legend_window.resizable(False, False)
        
        ttk.Label(legend_window, text="Уровни риска активности клещей", 
                 font=('Arial', 12, 'bold')).pack(pady=10)
        
        legend_frame = ttk.Frame(legend_window)
        legend_frame.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)
        
        risk_levels = [
            ("Низкий", "#00ff00", "Менее 50 случаев"),
            ("Умеренный", "#ffff00", "50-99 случаев"),
            ("Высокий", "#ff9900", "100-149 случаев"),
            ("Очень высокий", "#ff0000", "150+ случаев")
        ]
        
        for level, color, desc in risk_levels:
            frame = ttk.Frame(legend_frame)
            frame.pack(fill=tk.X, pady=2)
            
            canvas = tk.Canvas(frame, width=20, height=20, bg=self.bg_color, highlightthickness=0)
            canvas.create_rectangle(2, 2, 18, 18, fill=color, outline="white")
            canvas.pack(side=tk.LEFT, padx=5)
            
            ttk.Label(frame, text=f"{level}: {desc}").pack(side=tk.LEFT, anchor=tk.W)

    def toggle_sound(self):
        """Включение/выключение звука"""
        self.sound_enabled = not self.sound_enabled
        if self.sound_enabled:
            self.play_background_sound()
            self.sound_btn.config(text="🔊 Звук Вкл")
        else:
            pygame.mixer.stop()
            self.sound_btn.config(text="🔇 Звук Выкл")

    def apply_date_filter(self):
        """Применение фильтра по датам"""
        start_date = self.start_date.get_date()
        end_date = self.end_date.get_date()
        
        if start_date > end_date:
            messagebox.showerror("Ошибка", "Начальная дата не может быть позже конечной")
            return
        
        filtered_data = []
        for src in self.data.get('sources', []):
            src_date = src['date']
            if isinstance(src_date, str):
                try:
                    src_date = datetime.strptime(src_date, '%d.%m.%Y').date()
                except ValueError:
                    continue
            
            if start_date <= src_date <= end_date:
                filtered_data.append(src)
        
        if not filtered_data:
            messagebox.showinfo("Информация", "Нет данных за выбранный период")
            return
        
        filtered_data.sort(key=lambda x: x['date'] if isinstance(x['date'], date) else datetime.strptime(x['date'], '%d.%m.%Y').date())
        
        self.update_filtered_graph(filtered_data, start_date, end_date)
        self.notebook.select(0)

    def update_filtered_graph(self, data, start_date, end_date):
        """Обновление графика с отфильтрованными данными"""
        # Создаем DataFrame для удобной обработки
        df = pd.DataFrame(data)
        
        # Группируем по неделям
        df['date'] = pd.to_datetime(df['date'])
        df['year_week'] = df['date'].dt.strftime('%Y-%U')
        
        weekly_data = df.groupby('year_week').agg({
            'cases': 'sum',
            'date': ['min', 'max']
        }).reset_index()
        
        weekly_data.columns = ['year_week', 'cases', 'start_date', 'end_date']
        
        # Сортируем по дате
        weekly_data = weekly_data.sort_values('start_date')
        
        # Подготавливаем данные для графика
        weeks = []
        cases = []
        colors = []
        
        for _, row in weekly_data.iterrows():
            week_label = f"{row['start_date'].strftime('%d.%m')}-{row['end_date'].strftime('%d.%m')}"
            weeks.append(week_label)
            cases.append(row['cases'])
            risk_level = self.calculate_risk_level(row['cases'])
            colors.append(self.get_risk_color(risk_level))
        
        # Очищаем и обновляем график
        self.ax.clear()
        
        # Гистограмма с настройками
        bars = self.ax.bar(weeks, cases, color=colors, width=0.7, edgecolor='white', linewidth=1)
        
        # Настройки осей
        self.ax.set_ylabel('Количество обращений', color=self.text_color, fontsize=10)
        self.ax.set_xlabel('Период (неделя)', color=self.text_color, fontsize=10)
        self.ax.set_title(
            f'Активность клещей {start_date.strftime("%d.%m.%Y")}-{end_date.strftime("%d.%m.%Y")}', 
            color=self.highlight_color, fontsize=12, pad=20
        )
        
        # Поворачиваем подписи на оси X для лучшей читаемости
        plt.setp(self.ax.get_xticklabels(), rotation=45, ha='right', rotation_mode='anchor')
        
        # Добавляем значения на столбцы
        for bar in bars:
            height = bar.get_height()
            self.ax.text(
                bar.get_x() + bar.get_width()/2., 
                height + 0.5,
                f'{int(height)}',
                ha='center', va='bottom', 
                color='#ffffff',
                fontsize=8
            )
        
        # Настраиваем сетку
        self.ax.grid(True, linestyle='--', alpha=0.7, color='#444444')
        
        # Обновляем холст
        self.canvas_graph.draw()
        
        # Обновляем данные о текущей и предыдущей неделе
        if len(weekly_data) >= 2:
            last_week = weekly_data.iloc[-1]
            prev_week = weekly_data.iloc[-2]
            
            self.current_data.config(text=f"{last_week['cases']} случаев\n"
                                      f"({last_week['start_date'].strftime('%d.%m')}-{last_week['end_date'].strftime('%d.%m')})")
            self.prev_data.config(text=f"{prev_week['cases']} случаев\n"
                                    f"({prev_week['start_date'].strftime('%d.%m')}-{prev_week['end_date'].strftime('%d.%m')})")
            
            self.current_risk.config(
                text=f"Риск: {self.calculate_risk_level(last_week['cases'])}",
                foreground=self.get_risk_color(self.calculate_risk_level(last_week['cases']))
            )
            self.prev_risk.config(
                text=f"Риск: {self.calculate_risk_level(prev_week['cases'])}",
                foreground=self.get_risk_color(self.calculate_risk_level(prev_week['cases']))
            )

    def parse_rospotrebnadzor(self):
        """Парсинг данных с сайта Роспотребнадзора"""
        web_data = self.parse_web_data()
        rss_data = self.parse_rss_feed()
        telegram_data = self.parse_telegram()
        
        combined_results = []
        if web_data:
            combined_results.extend(web_data)
        if rss_data:
            combined_results.extend(rss_data)
        if telegram_data:
            combined_results.extend(telegram_data)
        
        if combined_results:
            combined_results.sort(key=lambda x: x['date'], reverse=True)
            
            # Сохраняем в Excel
            self.save_to_excel(combined_results)
            
            # Загружаем все данные из Excel (включая новые)
            all_data = self.load_from_excel()
            
            current_week_data = self.find_week_data(all_data, 0)
            previous_week_data = self.find_week_data(all_data, 1)
            
            max_sources = self.config.get('graph', {}).get('filtered_max_items', 100)
            return {
                'current_week': current_week_data,
                'previous_week': previous_week_data,
                'sources': all_data[:max_sources]
            }
        return None

    def parse_telegram(self):
        """Парсинг данных из Telegram-канала"""
        try:
            telegram_config = self.config.get('parsing', {}).get('sources', {}).get('telegram', {})
            url = telegram_config.get('url', 'https://t.me/s/tu_ymen72')
            max_items = telegram_config.get('max_items', 50)
            
            ua = UserAgent()
            headers = {
                'User-Agent': ua.random,
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
            }
            
            self.logger.info(f"Парсинг Telegram: {url}")
            response = self.make_request_with_retry(url, headers)
            
            if not response:
                return None
            
            soup = BeautifulSoup(response.text, 'html.parser')
            results = []
            
            # Ищем сообщения, содержащие информацию о клещах
            messages = soup.find_all('div', class_='tgme_widget_message')
            
            for message in messages[:max_items]:
                try:
                    # Пропускаем сообщения без текста
                    if not message.find('div', class_='tgme_widget_message_text'):
                        continue
                        
                    text = message.find('div', class_='tgme_widget_message_text').get_text('\n', strip=True)
                    
                    # Ищем только сообщения, связанные с клещами
                    if not any(word in text.lower() for word in ['клещ', 'укус', 'энцефалит', 'присасыван']):
                        continue
                    
                    # Извлекаем дату сообщения
                    time_tag = message.find('time', class_='time')
                    if not time_tag or not time_tag.has_attr('datetime'):
                        continue
                        
                    message_date = date_parser.parse(time_tag['datetime']).date()
                    
                    # Извлекаем количество случаев
                    cases = self.extract_case_number(text)
                    if not cases:
                        cases = 0
                    
                    # Формируем результат
                    results.append({
                        'date': message_date,
                        'cases': cases,
                        'title': text[:50] + "..." if len(text) > 50 else text,
                        'content': text[:200] + "..." if len(text) > 200 else text,
                        'url': url,
                        'source': 'Telegram (Тюмень 72)'
                    })
                except Exception as e:
                    self.logger.debug(f"Ошибка обработки сообщения Telegram: {str(e)}")
                    continue
            
            self.logger.info(f"Получено {len(results)} записей из Telegram")
            return results
            
        except Exception as e:
            self.logger.error(f"Ошибка при парсинге Telegram: {str(e)}")
            return None

    def find_week_data(self, data, weeks_ago):
        """Находит данные за указанное количество недель назад"""
        if not data:
            return {"cases": 0, "date": datetime.now().date(), "risk_level": "Нет данных"}
            
        target_date = datetime.now().date() - timedelta(weeks=weeks_ago)
        
        for item in data:
            item_date = item['date'] if isinstance(item['date'], date) else datetime.strptime(item['date'], '%d.%m.%Y').date()
            if item_date <= target_date:
                return {
                    'cases': item['cases'],
                    'date': item_date,
                    'risk_level': self.calculate_risk_level(item['cases'])
                }
        
        return {
            'cases': data[-1]['cases'],
            'date': data[-1]['date'],
            'risk_level': self.calculate_risk_level(data[-1]['cases'])
        }

    def make_request_with_retry(self, url, headers, max_retries=3, delay=2):
        """Выполнение HTTP запроса с повторами при ошибках"""
        parsing_config = self.config.get('parsing', {})
        max_retries = parsing_config.get('retry_count', max_retries)
        delay = parsing_config.get('retry_delay', delay)
        timeout = parsing_config.get('timeout', 15)
        
        for attempt in range(max_retries):
            try:
                response = requests.get(url, headers=headers, timeout=timeout)
                response.raise_for_status()
                return response
            except requests.exceptions.RequestException as e:
                self.logger.warning(f"Попытка {attempt + 1}/{max_retries} не удалась для {url}: {str(e)}")
                if attempt < max_retries - 1:
                    time.sleep(delay)
                else:
                    self.logger.error(f"Все попытки исчерпаны для {url}")
                    raise
        return None
    
    def parse_web_data(self):
        """Парсинг данных с веб-сайта"""
        try:
            ua = UserAgent()
            headers = {
                'User-Agent': ua.random,
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'ru-RU,ru;q=0.8,en-US;q=0.5,en;q=0.3',
            }
            
            web_config = self.config.get('parsing', {}).get('sources', {}).get('web', {})
            base_url = web_config.get('base_url', 'https://72.rospotrebnadzor.ru')
            search_url = web_config.get('search_url', f"{base_url}/search/?q=%D0%BA%D0%BB%D0%B5%D1%89%D0%B8")
            max_items = web_config.get('max_items', 100)
            
            self.logger.info(f"Парсинг веб-сайта: {search_url}")
            response = self.make_request_with_retry(search_url, headers)
            
            if not response:
                return None
            
            soup = BeautifulSoup(response.text, 'html.parser')
            results = []
            
            news_items = soup.find_all('div', class_='search-item')[:max_items]
            
            for item in news_items:
                try:
                    title = item.find('a', class_='search-title').text.strip()
                    date_text = item.find('div', class_='search-date').text.strip()
                    content = item.find('div', class_='search-text').text.strip()
                    
                    date_match = re.search(r'\d{2}\.\d{2}\.\d{4}', date_text)
                    if not date_match:
                        continue
                        
                    date = datetime.strptime(date_match.group(), '%d.%m.%Y').date()
                    
                    cases = self.extract_case_number(title + " " + content)
                    if cases:
                        results.append({
                            'date': date,
                            'cases': cases,
                            'title': title,
                            'content': content[:200] + "...",
                            'url': base_url + item.find('a')['href'],
                            'source': 'Роспотребнадзор (веб)'
                        })
                except Exception as e:
                    self.logger.debug(f"Ошибка обработки новости: {str(e)}")
                    continue
            
            self.logger.info(f"Получено {len(results)} записей с веб-сайта")
            return results
            
        except Exception as e:
            self.logger.error(f"Ошибка при парсинге веб-сайта: {str(e)}")
            return None

    def parse_rss_feed(self):
        """Парсинг RSS-ленты"""
        try:
            web_config = self.config.get('parsing', {}).get('sources', {}).get('web', {})
            rss_url = web_config.get('rss_url', 'https://72.rospotrebnadzor.ru/rss/')
            max_items = web_config.get('max_items', 100)
            
            self.logger.info(f"Парсинг RSS-ленты: {rss_url}")
            feed = feedparser.parse(rss_url)
            results = []
            
            for entry in feed.entries[:max_items]:
                try:
                    date = datetime(*entry.published_parsed[:6]).date()
                    
                    if any(word in entry.title.lower() or word in entry.description.lower() 
                          for word in ['клещ', 'укус', 'энцефалит', 'присасыван']):
                        
                        cases = self.extract_case_number(entry.title + " " + entry.description)
                        if cases:
                            results.append({
                                'date': date,
                                'cases': cases,
                                'title': entry.title,
                                'content': entry.description[:200] + "...",
                                'url': entry.link,
                                'source': 'Роспотребнадзор (RSS)'
                            })
                except Exception as e:
                    self.logger.debug(f"Ошибка обработки RSS-записи: {str(e)}")
                    continue
            
            self.logger.info(f"Получено {len(results)} записей из RSS")
            return results
            
        except Exception as e:
            self.logger.error(f"Ошибка при парсинге RSS: {str(e)}")
            return None

    def extract_case_number(self, text):
        """Извлекает количество случаев из текста"""
        patterns = [
            r'зарегистрировано\D*(\d+)\D*обращ',
            r'выявлено\D*(\d+)\D*случа',
            r'(\d+)\D*укус',
            r'клещ\D*(\d+)',
            r'(\d+)\s*(?:случа[ея]в|обращени[ий])'
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                try:
                    return int(match.group(1))
                except ValueError:
                    continue
        return 0  # Возвращаем 0 вместо None для лучшей обработки в Excel

    def update_data(self):
        """Основной метод обновления данных"""
        try:
            self.logger.info("Начало обновления данных")
            web_data = self.parse_rospotrebnadzor()
            
            if web_data:
                self.process_new_data(web_data)
                self.logger.info("Данные успешно обновлены")
                messagebox.showinfo("Успех", "Данные успешно обновлены!")
                self.play_sound(self.update_sound)
            else:
                raise ValueError("Не удалось получить данные с сайтов")
                
        except Exception as e:
            self.logger.error(f"Ошибка при обновлении: {str(e)}", exc_info=True)
            self.load_backup_data()
            messagebox.showwarning("Внимание", 
                f"Не удалось загрузить новые данные: {str(e)}\nИспользуются сохранённые данные.")

    def update_data_threaded(self):
        """Запуск обновления в отдельном потоке"""
        self.update_btn.config(state=tk.DISABLED)
        self.update_label.config(text="Обновление данных...")
        thread = threading.Thread(target=self.update_data)
        thread.daemon = True
        thread.start()
        self.root.after(100, self.check_thread, thread)

    def check_thread(self, thread):
        """Проверка завершения потока"""
        if thread.is_alive():
            self.root.after(100, self.check_thread, thread)
        else:
            self.update_btn.config(state=tk.NORMAL)
            self.update_label.config(text=f"Обновлено: {self.data.get('last_update', 'никогда')}")

    def process_new_data(self, new_data):
        """Обработка новых данных"""
        self.data = {
            "current_week": {
                "cases": new_data['current_week']['cases'],
                "date": new_data['current_week']['date'].strftime('%d.%m.%Y'),
                "risk_level": self.calculate_risk_level(new_data['current_week']['cases'])
            },
            "previous_week": {
                "cases": new_data['previous_week']['cases'],
                "date": new_data['previous_week']['date'].strftime('%d.%m.%Y'),
                "risk_level": self.calculate_risk_level(new_data['previous_week']['cases'])
            },
            "sources": new_data['sources'],
            "last_update": datetime.now().strftime("%d.%m.%Y %H:%M:%S")
        }
        
        self.update_ui()
        self.save_data()

    def update_ui(self):
        """Обновление интерфейса"""
        current = self.data['current_week']
        prev = self.data['previous_week']
        
        self.current_data.config(text=f"{current['cases']} случаев\n({current['date']})")
        self.prev_data.config(text=f"{prev['cases']} случаев\n({prev['date']})")
        
        self.current_risk.config(
            text=f"Риск: {current['risk_level']}",
            foreground=self.get_risk_color(current['risk_level'])
        )
        self.prev_risk.config(
            text=f"Риск: {prev['risk_level']}",
            foreground=self.get_risk_color(prev['risk_level'])
        )
        
        self.update_graph()
        self.update_sources()
        self.update_label.config(text=f"Обновлено: {self.data.get('last_update', 'никогда')}")

    def update_graph(self):
        """Обновление графика с данными из Excel"""
        try:
            # Загружаем все данные из Excel
            all_data = self.load_from_excel()
            
            if not all_data:
                return
                
            # Создаем DataFrame для анализа
            df = pd.DataFrame(all_data)
            df['date'] = pd.to_datetime(df['date'])
            
            # Группируем по неделям
            df['year_week'] = df['date'].dt.strftime('%Y-%U')
            weekly_data = df.groupby('year_week').agg({
                'cases': 'sum',
                'date': ['min', 'max']
            }).reset_index()
            
            weekly_data.columns = ['year_week', 'cases', 'start_date', 'end_date']
            weekly_data = weekly_data.sort_values('start_date')
            
            # Берем последние N недель для графика из конфига
            weeks_to_show = self.config.get('graph', {}).get('weeks_to_show', 8)
            weekly_data = weekly_data.tail(weeks_to_show)
            
            # Подготовка данных для графика
            weeks = []
            cases = []
            colors = []
            
            for _, row in weekly_data.iterrows():
                week_label = f"{row['start_date'].strftime('%d.%m')}-{row['end_date'].strftime('%d.%m')}"
                weeks.append(week_label)
                cases.append(row['cases'])
                colors.append(self.get_risk_color(self.calculate_risk_level(row['cases'])))
            
            # Очищаем и обновляем график
            self.ax.clear()
            
            # Гистограмма с улучшенным дизайном
            bars = self.ax.bar(
                weeks, cases, 
                color=colors,
                width=0.6,
                edgecolor='white',
                linewidth=1,
                alpha=0.8
            )
            
            # Настройки осей и заголовка
            self.ax.set_title(
                'Динамика активности клещей (последние 8 недель)',
                color=self.highlight_color,
                fontsize=12,
                pad=20
            )
            self.ax.set_ylabel('Количество обращений', color=self.text_color)
            self.ax.set_xlabel('Неделя', color=self.text_color)
            
            # Поворот подписей на оси X
            plt.setp(self.ax.get_xticklabels(), rotation=45, ha='right')
            
            # Добавление значений на столбцы
            for bar in bars:
                height = bar.get_height()
                self.ax.text(
                    bar.get_x() + bar.get_width()/2.,
                    height + 0.5,
                    f'{int(height)}',
                    ha='center',
                    va='bottom',
                    color='white',
                    fontsize=9
                )
            
            # Настройка сетки
            self.ax.grid(True, linestyle=':', color='#555555', alpha=0.5)
            
            # Обновляем холст
            self.canvas_graph.draw()
            
        except Exception as e:
            self.logger.error(f"Ошибка при обновлении графика: {str(e)}", exc_info=True)

    def update_sources(self):
        """Обновление списка источников"""
        self.sources_text.config(state=tk.NORMAL)
        self.sources_text.delete(1.0, tk.END)
        
        # Показываем только последние 20 записей
        for src in self.data.get('sources', [])[:20]:
            date_str = src['date'].strftime('%d.%m.%Y') if isinstance(src['date'], date) else src['date']
            self.sources_text.insert(tk.END, 
                                   f"Дата: {date_str}\n"
                                   f"Случаев: {src['cases']}\n"
                                   f"Риск: {src.get('risk_level', 'Нет данных')}\n"
                                   f"Источник: {src.get('source', 'Неизвестно')}\n"
                                   f"Заголовок: {src.get('title', '')}\n"
                                   f"Ссылка: {src.get('url', '')}\n"
                                   f"{'-'*50}\n")
        
        self.sources_text.config(state=tk.DISABLED)

    def calculate_risk_level(self, cases):
        """Определение уровня риска на основе конфигурации"""
        if not isinstance(cases, int) or cases == 0:
            return "Нет данных"
        
        risk_config = self.config.get('risk_levels', {})
        thresholds = {
            'low': risk_config.get('low', {}).get('threshold', 50),
            'moderate': risk_config.get('moderate', {}).get('threshold', 100),
            'high': risk_config.get('high', {}).get('threshold', 150),
            'very_high': risk_config.get('very_high', {}).get('threshold', 999999)
        }
        
        if cases < thresholds['low']:
            return "Низкий"
        elif cases < thresholds['moderate']:
            return "Умеренный"
        elif cases < thresholds['high']:
            return "Высокий"
        else:
            return "Очень высокий"

    def get_risk_color(self, risk_level):
        """Цвет для уровня риска из конфигурации"""
        risk_config = self.config.get('risk_levels', {})
        colors = {
            "Низкий": risk_config.get('low', {}).get('color', '#00ff00'),
            "Умеренный": risk_config.get('moderate', {}).get('color', '#ffff00'),
            "Высокий": risk_config.get('high', {}).get('color', '#ff9900'),
            "Очень высокий": risk_config.get('very_high', {}).get('color', '#ff0000'),
            "Нет данных": "#aaaaaa"
        }
        return colors.get(risk_level, "#aaaaaa")

    def save_data(self):
        """Сохранение данных в файл JSON"""
        try:
            data_to_save = self.data.copy()
            for src in data_to_save['sources']:
                if isinstance(src['date'], date):
                    src['date'] = src['date'].strftime('%d.%m.%Y')
            
            with open(self.data_file, 'w', encoding='utf-8') as f:
                json.dump(data_to_save, f, ensure_ascii=False, indent=2)
            self.logger.debug("Данные сохранены в JSON")
        except Exception as e:
            self.logger.error(f"Ошибка при сохранении JSON: {str(e)}")

    def load_data(self):
        """Загрузка данных из файла JSON и Excel"""
        try:
            # Загружаем из JSON
            if os.path.exists(self.data_file):
                with open(self.data_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    
                    for src in data.get('sources', []):
                        if 'date' in src and isinstance(src['date'], str):
                            try:
                                src['date'] = datetime.strptime(src['date'], '%d.%m.%Y').date()
                            except ValueError:
                                continue
                    
                    self.data = data
            
            # Дополняем данными из Excel
            excel_data = self.load_from_excel()
            if excel_data:
                if 'sources' not in self.data:
                    self.data['sources'] = []
                
                # Объединяем источники, избегая дубликатов
                existing_sources = {(src['date'], src['source'], src['title']) for src in self.data['sources']}
                for src in excel_data:
                    key = (src['date'], src['source'], src['title'])
                    if key not in existing_sources:
                        self.data['sources'].append(src)
                
                # Сортируем по дате
                self.data['sources'].sort(key=lambda x: x['date'], reverse=True)
                
                # Обновляем текущую и предыдущую неделю
                if len(self.data['sources']) >= 1:
                    self.data['current_week'] = {
                        'cases': self.data['sources'][0]['cases'],
                        'date': self.data['sources'][0]['date'],
                        'risk_level': self.calculate_risk_level(self.data['sources'][0]['cases'])
                    }
                
                if len(self.data['sources']) >= 2:
                    self.data['previous_week'] = {
                        'cases': self.data['sources'][1]['cases'],
                        'date': self.data['sources'][1]['date'],
                        'risk_level': self.calculate_risk_level(self.data['sources'][1]['cases'])
                    }
            
            self.update_ui()
            self.logger.info("Данные загружены успешно")
            
        except (FileNotFoundError, json.JSONDecodeError) as e:
            self.logger.warning(f"Ошибка загрузки: {str(e)}, используются значения по умолчанию")
            self.data = self.default_data.copy()

    def load_backup_data(self):
        """Загрузка резервных данных"""
        self.load_data()

    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    logger = setup_logger()
    try:
        logger.info("Запуск приложения")
        root = tk.Tk()
        app = TickActivityMonitor(root)
        app.run()
    except Exception as e:
        logger.critical(f"Критическая ошибка: {str(e)}", exc_info=True)
        print(f"Произошла критическая ошибка: {str(e)}")
        sys.exit(1)